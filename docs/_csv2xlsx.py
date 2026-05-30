import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

SRC = '/Users/vodinhkhang/Documents/HaHaGoWeb/docs/dev-checklist.csv'
DST = '/Users/vodinhkhang/Documents/HaHaGoWeb/docs/dev-checklist.xlsx'

wb = Workbook()
ws = wb.active
ws.title = 'Checklist'

with open(SRC, encoding='utf-8') as f:
    rows = list(csv.reader(f))

for r in rows:
    ws.append(r)

# Header style
header_fill = PatternFill('solid', fgColor='1F4E78')
header_font = Font(bold=True, color='FFFFFF', size=11)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Column widths
widths = [8, 18, 26, 28, 40, 10, 35, 12, 14, 12, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Wrap text + vertical align top for all data
thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(widths)):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = border

# Row height auto-ish
for i in range(2, ws.max_row + 1):
    ws.row_dimensions[i].height = 45

# Freeze header + first col
ws.freeze_panes = 'C2'

# Priority color coding (col F = 6)
p0_fill = PatternFill('solid', fgColor='F8CBAD')  # đỏ nhạt
p1_fill = PatternFill('solid', fgColor='FFE699')  # vàng
p2_fill = PatternFill('solid', fgColor='C6EFCE')  # xanh
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    pcell = row[5]  # Priority col
    v = (pcell.value or '').strip()
    if v == 'P0': pcell.fill = p0_fill
    elif v == 'P1': pcell.fill = p1_fill
    elif v == 'P2': pcell.fill = p2_fill
    pcell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    pcell.font = Font(bold=True)

# Status color coding (col H = 8)
open_fill = PatternFill('solid', fgColor='FCE4D6')
pass_fill = PatternFill('solid', fgColor='C6EFCE')
fail_fill = PatternFill('solid', fgColor='FFC7CE')
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    scell = row[7]
    v = (scell.value or '').strip().lower()
    if v == 'open': scell.fill = open_fill
    elif v == 'pass': scell.fill = pass_fill
    elif v == 'fail': scell.fill = fail_fill
    scell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Data validation: Trạng thái + Priority dropdown
from openpyxl.worksheet.datavalidation import DataValidation
dv_status = DataValidation(type='list', formula1='"Open,Pass,Fail,Blocked,Skipped"', allow_blank=True)
dv_status.add(f'H2:H{ws.max_row}')
ws.add_data_validation(dv_status)

dv_pri = DataValidation(type='list', formula1='"P0,P1,P2"', allow_blank=True)
dv_pri.add(f'F2:F{ws.max_row}')
ws.add_data_validation(dv_pri)

# AutoFilter
ws.auto_filter.ref = ws.dimensions

# --- Sheet 2: Summary by Priority ---
ws2 = wb.create_sheet('Summary')
ws2.append(['Priority', 'Tổng', 'Open', 'Pass', 'Fail'])
for c in ws2[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center')

max_row = ws.max_row
for i, p in enumerate(['P0', 'P1', 'P2'], 2):
    ws2.cell(i, 1, p).font = Font(bold=True)
    ws2.cell(i, 2, f'=COUNTIF(Checklist!F2:F{max_row},"{p}")')
    ws2.cell(i, 3, f'=COUNTIFS(Checklist!F2:F{max_row},"{p}",Checklist!H2:H{max_row},"Open")')
    ws2.cell(i, 4, f'=COUNTIFS(Checklist!F2:F{max_row},"{p}",Checklist!H2:H{max_row},"Pass")')
    ws2.cell(i, 5, f'=COUNTIFS(Checklist!F2:F{max_row},"{p}",Checklist!H2:H{max_row},"Fail")')

ws2.append(['TỔNG', f'=SUM(B2:B4)', f'=SUM(C2:C4)', f'=SUM(D2:D4)', f'=SUM(E2:E4)'])
for c in ws2[5]:
    c.font = Font(bold=True)
    c.fill = PatternFill('solid', fgColor='D9E1F2')

for col, w in enumerate([12, 10, 10, 10, 10], 1):
    ws2.column_dimensions[get_column_letter(col)].width = w

# --- Sheet 3: Spec headers (link reference) ---
ws3 = wb.create_sheet('Spec_Headers')
ws3.append(['Xem chi tiết spec tại file:', 'docs/table-headers-spec.md'])
ws3['A1'].font = Font(bold=True, size=12)
ws3['B1'].font = Font(italic=True, color='0563C1')
ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 40

wb.save(DST)
print(f'Saved: {DST}')
